"""
Backfill the local ``paid_date`` field on existing contracts from a spreadsheet.

The spreadsheet is expected to have two columns with a header row:

    YTD Paid Date | Contract

Matching is by ``salesorder_number`` (stored as a string in Quern). Where a
contract number appears on multiple rows (installment payments), the most
recent date wins. Contracts not present in Quern are skipped and reported.
``paid_date`` is a Quern-local field, so this touches SQLite only — no Zoho
API calls are made.

Usage (from the project root, inside the app venv):

    python import_payments.py <path-to-xlsx>            # apply changes
    python import_payments.py <path-to-xlsx> --dry-run  # report only, no writes

On Azure, run over SSH (Kudu) with the app venv active so DATABASE_URL points
at /mnt/quern-data/quern.db. Locally it falls back to the local quern.db.
"""
import sys
from datetime import datetime

from openpyxl import load_workbook

from app import app
from core.models import db, Contract


def _normalize_contract_number(value):
    """Return the spreadsheet contract cell as a trimmed string.

    Integer contract numbers (e.g. 2512072) arrive as ints and must be
    stringified to match the String salesorder_number column. Prefixed
    numbers (e.g. 'SBC2507019') arrive as strings already.
    """
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _normalize_date(value):
    """Return an ISO YYYY-MM-DD string, or '' if the cell can't be parsed."""
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    # Fallback: accept strings already in ISO-ish form.
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return ''


def load_latest_dates(xlsx_path):
    """Read the spreadsheet and collapse to {contract_number: latest_date}.

    Returns (latest_by_contract, stats) where stats records rows read and any
    rows skipped for unparseable data.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    latest = {}
    rows_read = 0
    skipped_bad_contract = 0
    skipped_bad_date = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header row
        if row is None or len(row) < 2:
            continue
        rows_read += 1

        contract = _normalize_contract_number(row[1])
        date_str = _normalize_date(row[0])

        if not contract:
            skipped_bad_contract += 1
            continue
        if not date_str:
            skipped_bad_date += 1
            continue

        # ISO dates sort lexically, so max() gives the most recent date.
        existing = latest.get(contract)
        if existing is None or date_str > existing:
            latest[contract] = date_str

    wb.close()
    stats = {
        'rows_read': rows_read,
        'skipped_bad_contract': skipped_bad_contract,
        'skipped_bad_date': skipped_bad_date,
        'unique_contracts': len(latest),
    }
    return latest, stats


def run(xlsx_path, dry_run=False):
    latest_by_contract, stats = load_latest_dates(xlsx_path)

    print('── Spreadsheet ─────────────────────────────')
    print(f'  Rows read (excl. header):     {stats["rows_read"]}')
    print(f'  Unique contract numbers:      {stats["unique_contracts"]}')
    if stats['skipped_bad_contract']:
        print(f'  Skipped (blank contract):     {stats["skipped_bad_contract"]}')
    if stats['skipped_bad_date']:
        print(f'  Skipped (unparseable date):   {stats["skipped_bad_date"]}')
    print()

    updated = 0
    unchanged = 0
    missing = []

    with app.app_context():
        for contract_number, date_str in latest_by_contract.items():
            contract = Contract.query.filter_by(
                salesorder_number=contract_number
            ).first()

            if not contract:
                missing.append(contract_number)
                continue

            if contract.paid_date == date_str:
                unchanged += 1
                continue

            if not dry_run:
                contract.paid_date = date_str
            updated += 1

        if not dry_run:
            db.session.commit()

    print('── Result ──────────────────────────────────')
    verb = 'Would update' if dry_run else 'Updated'
    print(f'  {verb}:                     {updated}')
    print(f'  Already correct (skipped):    {unchanged}')
    print(f'  Not found in Quern (skipped): {len(missing)}')
    print()

    if missing:
        print('── Contract numbers not found in Quern ─────')
        for c in missing:
            print(f'  {c}')
        print()

    if dry_run:
        print('DRY RUN — no changes were written.')
    else:
        print('Done. Changes committed.')


def main():
    args = [a for a in sys.argv[1:]]
    dry_run = '--dry-run' in args
    positional = [a for a in args if not a.startswith('--')]

    if len(positional) != 1:
        print('Usage: python import_payments.py <path-to-xlsx> [--dry-run]')
        sys.exit(1)

    run(positional[0], dry_run=dry_run)


if __name__ == '__main__':
    main()