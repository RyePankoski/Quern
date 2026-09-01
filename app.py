# region Imports
import os
from datetime import timezone
from flask import Flask, render_template, request, redirect, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.cron import CronTrigger
import atexit
from datetime import datetime

from core import zoho

from core.pdf import generate_contract_pdf, generate_contract_docx
from core.tasks import generate_tasks, check_task_reactivity
from core.zoho import get_sales_orders, sync_employees, sync_items, sync_customers
from core.zoho import sync_customers_page, sync_contracts_page
from core.powerbi_export import export_contracts_to_excel

from functions import *

# endregion

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///quern.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = os.getenv("SECRET_KEY")
db.init_app(app)
migrate = Migrate(app, db, render_as_batch=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ── Scheduler Setup ────────────────────────────────────────────────────────────
scheduler = BackgroundScheduler()
scheduler_log = []


def log_scheduler_message(msg):
    """Log a message with timestamp to the scheduler log."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_entry = f'[{timestamp}] {msg}'
    scheduler_log.append(log_entry)
    # Keep last 100 messages
    if len(scheduler_log) > 100:
        scheduler_log.pop(0)
    print(log_entry)


def scheduler_activate_tasks():
    """Check and activate tasks whose scheduling time has arrived."""
    try:
        with app.app_context():
            from core.models import Task
            from datetime import datetime, timedelta

            now = datetime.utcnow()
            activated = 0

            # Find tasks that should be visible but aren't yet
            hidden_tasks = Task.query.filter_by(is_visible=False).all()

            for task in hidden_tasks:
                should_show = False
                reason = ''

                # Check if scheduled_date has passed
                if task.scheduled_date and task.scheduled_date <= now:
                    should_show = True
                    reason = f'scheduled date {task.scheduled_date.strftime("%Y-%m-%d %H:%M:%S")} reached'

                # Check if delay_minutes has elapsed
                if not should_show and task.delay_minutes and task.created_at:
                    show_time = task.created_at + timedelta(minutes=task.delay_minutes)
                    if show_time <= now:
                        should_show = True
                        reason = f'delay of {task.delay_minutes} minutes elapsed'

                if should_show:
                    task.is_visible = True
                    activated += 1
                    log_scheduler_message(f'Task {task.id} activated ({reason})')

            if activated > 0:
                db.session.commit()
    except Exception as e:
        log_scheduler_message(f'Error in activate_tasks: {str(e)}')


def scheduler_sync_items():
    """Daily sync of items from Zoho."""
    try:
        with app.app_context():
            log_scheduler_message('Starting daily items sync...')
            sync_items()
            log_scheduler_message('Daily items sync completed')
    except Exception as e:
        log_scheduler_message(f'Error in sync_items: {str(e)}')


def scheduler_sync_counterparties():
    """Daily sync of counterparties (customers) from Zoho."""
    try:
        with app.app_context():
            log_scheduler_message('Starting daily counterparties sync...')
            sync_customers()
            log_scheduler_message('Daily counterparties sync completed')
    except Exception as e:
        log_scheduler_message(f'Error in sync_counterparties: {str(e)}')


def scheduler_sync_contracts():
    """Incremental sync of changed contracts from Zoho."""
    try:
        with app.app_context():
            log_scheduler_message('Starting incremental contracts sync...')
            page = 1
            total = 0
            while True:
                result = zoho.incremental_sync_page(page)
                total += result['count']
                log_scheduler_message(f'  Synced page {page}: {result["count"]} contracts')
                if not result['has_more']:
                    break
                page += 1
            log_scheduler_message(f'Incremental contracts sync completed ({total} total)')
    except Exception as e:
        log_scheduler_message(f'Error in sync_contracts: {str(e)}')


def scheduler_export_powerbi():
    """Export contracts to Excel for Power BI daily refresh."""
    try:
        with app.app_context():
            log_scheduler_message('Starting Power BI export...')
            output_path, result = export_contracts_to_excel()
            if output_path:
                log_scheduler_message(f'Power BI export completed: {result} contracts exported to {output_path}')
            else:
                log_scheduler_message(f'Power BI export failed: {result}')
    except Exception as e:
        log_scheduler_message(f'Error in export_powerbi: {str(e)}')


# Start scheduler on app startup
if not scheduler.running:
    scheduler.add_job(
        func=scheduler_activate_tasks,
        trigger=IntervalTrigger(minutes=1),  # Check every minute for scheduled tasks
        id='activate_tasks',
        name='Activate Scheduled Tasks',
        replace_existing=True
    )
    scheduler.add_job(
        func=scheduler_sync_items,
        trigger=CronTrigger(hour=0, minute=0),  # Every day at midnight UTC
        id='daily_sync_items',
        name='Daily Sync Items',
        replace_existing=True
    )
    scheduler.add_job(
        func=scheduler_sync_counterparties,
        trigger=CronTrigger(hour=0, minute=5),  # Every day at 12:05 AM UTC
        id='daily_sync_counterparties',
        name='Daily Sync Counterparties',
        replace_existing=True
    )
    scheduler.add_job(
        func=scheduler_sync_contracts,
        trigger=CronTrigger(hour=0, minute=10),  # Every day at 12:10 AM UTC
        id='daily_sync_contracts',
        name='Daily Sync Contracts',
        replace_existing=True
    )
    scheduler.add_job(
        func=scheduler_export_powerbi,
        trigger=CronTrigger(hour=0, minute=15),  # Every day at 12:15 AM UTC (after contracts sync)
        id='daily_export_powerbi',
        name='Daily Power BI Export',
        replace_existing=True
    )
    scheduler.start()
    log_scheduler_message('Scheduler started')

# Shut down scheduler on app exit
atexit.register(lambda: scheduler.shutdown() if scheduler.running else None)


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Admin access required.', 'danger')
            return redirect('/')
        return f(*args, **kwargs)

    return decorated


def _contracts_user_is_on():
    """Sales-order IDs where the current user is on the broker team or is the PIC.
    Used to let brokers see their own contracts regardless of nationality."""
    from sqlalchemy import func
    if current_user.is_admin:
        return set()
    email = (current_user.email or '').lower()
    if not email:
        return set()
    emp = Employee.query.filter(func.lower(Employee.email) == email).first()
    if not emp:
        return set()
    ids = {bc.books_sales_order_id
           for bc in BrokerCommission.query.filter_by(employee_id=emp.employee_id).all()}
    ids |= {c.salesorder_id
            for c in Contract.query.filter_by(pic_employee_id=emp.employee_id).all()}
    return ids


def _apply_contract_access(q):
    """Apply nationality-based access control to a Contract query.
    Non-admin users with a nationality see only contracts at their office,
    plus any contracts they're personally on the broker team for."""
    if not current_user.is_admin and current_user.nationality:
        own_ids = _contracts_user_is_on()
        if own_ids:
            from sqlalchemy import or_
            q = q.filter(or_(
                Contract.location_name == current_user.nationality,
                Contract.salesorder_id.in_(own_ids)
            ))
        else:
            q = q.filter(Contract.location_name == current_user.nationality)
    return q


# region Tabs/Pages
@app.route('/')
@login_required
def home():
    return render_template('home.html')


@app.route('/items')
@login_required
def items_view():
    items = Item.query.order_by(Item.item_name).all()
    return render_template('items.html', items=items)


@app.route('/items/<item_id>')
@login_required
def item_detail(item_id):
    item = Item.query.get(item_id)
    if not item:
        flash('Item not found.', 'danger')
        return redirect('/items')

    contracts_q = Contract.query.filter_by(item_id=item_id)
    contracts_q = _apply_contract_access(contracts_q)
    contracts_list = contracts_q.order_by(Contract.date.desc()).all()
    return render_template('item_detail.html', item=item, contracts=contracts_list)


@app.route('/counterparties')
@login_required
def counterparties():
    customers = Customer.query.order_by(Customer.customer_name).all()
    return render_template('counterparties.html', customers=customers)


@app.route('/counterparties/<customer_id>')
@login_required
def counterparty_detail(customer_id):
    customer = Customer.query.get(customer_id)
    if not customer:
        flash('Counterparty not found.', 'danger')
        return redirect('/counterparties')

    contracts_q = Contract.query.filter(
        db.or_(Contract.customer_id == customer_id, Contract.cf_buyer_id == customer_id)
    )
    contracts_q = _apply_contract_access(contracts_q)
    contracts_list = contracts_q.order_by(Contract.date.desc()).all()

    return render_template('counterparty_detail.html', customer=customer, contracts=contracts_list)


@app.route('/employees')
@login_required
def employees_view():
    employees = Employee.query.order_by(Employee.employee_name).all()
    return render_template('employees.html', employees=employees)


@app.route('/new_contract')
@login_required
def new_contract():
    items = Item.query.filter_by(is_active=True).order_by(Item.item_name).all()
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.customer_name).all()
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.employee_name).all()
    locations = zoho.get_locations()
    prefill = None
    prefill_brokers = []
    duplicate_id = request.args.get('duplicate')

    if duplicate_id:
        source = Contract.query.get(duplicate_id)
        if source:
            prefill = zoho.contract_to_form_data_local(source)
            prefill_brokers = BrokerCommission.query.filter_by(books_sales_order_id=duplicate_id).all()
            if not prefill_brokers and source.salesperson_id:
                fallback_emp = Employee.query.filter_by(salesperson_id=source.salesperson_id).first()
                if fallback_emp:
                    prefill_brokers = [type('obj', (object,), {
                        'employee_id': fallback_emp.employee_id,
                        'percentage': 100.0
                    })()]

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('new_contract.html', items=items, customers=customers, employees=employees, prefill=prefill,
                           prefill_brokers=prefill_brokers, locations=locations, today=today)


@app.route('/bulk_edit')
@login_required
def bulk_edit():
    q = Contract.query

    # Nationality-based access control
    q = _apply_contract_access(q)

    # Text filters
    number = request.args.get('number', '').strip()
    buyer = request.args.get('buyer', '').strip()
    seller = request.args.get('seller', '').strip()
    commodity = request.args.get('commodity', '').strip()
    broker = request.args.get('broker', '').strip()
    office = request.args.get('office', '').strip()
    status = request.args.get('status', '').strip()
    uom = request.args.get('uom', '').strip()
    transport = request.args.get('transport', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    ship_from = request.args.get('ship_from', '').strip()
    ship_to = request.args.get('ship_to', '').strip()
    universal = request.args.get('q', '').strip()
    exact = request.args.get('exact', '').strip()

    if number:    q = q.filter(Contract.salesorder_number.ilike(f'%{number}%'))
    if buyer:     q = q.filter(Contract.cf_buyer.ilike(f'%{buyer}%'))
    if seller:    q = q.filter(Contract.customer_name.ilike(f'%{seller}%'))
    if commodity: q = q.filter(Contract.item_name.ilike(f'%{commodity}%'))
    if broker:    q = q.filter(Contract.salesperson_name.ilike(f'%{broker}%'))
    if office:    q = q.filter(Contract.location_name.ilike(f'%{office}%'))
    if uom:       q = q.filter(Contract.cf_uom.ilike(f'%{uom}%'))
    if transport: q = q.filter(Contract.cf_trnspname == transport)
    if date_from: q = q.filter(Contract.date >= date_from)
    if date_to:   q = q.filter(Contract.date <= date_to)
    if ship_from: q = q.filter(Contract.shipment_date >= ship_from)
    if ship_to:   q = q.filter(Contract.shipment_date <= ship_to)
    if status == 'declined':
        q = q.filter(Contract.is_declined == True)  # noqa
    elif status:
        q = q.filter(Contract.status == status)

    if universal:
        from sqlalchemy import or_
        _ucols = [Contract.salesorder_number, Contract.cf_buyer, Contract.customer_name,
                  Contract.item_name, Contract.salesperson_name, Contract.location_name,
                  Contract.cf_vessel_name, Contract.cf_customer_ref, Contract.buyer_reference,
                  Contract.reference_number, Contract.cf_co_broker, Contract.cf_split_broker,
                  Contract.cf_origin_location, Contract.cf_trnspname, Contract.cf_uom]
        if exact:
            q = q.filter(or_(*[c == universal for c in _ucols]))
        else:
            q = q.filter(or_(*[c.ilike(f'%{universal}%') for c in _ucols]))

    filters = dict(number=number, buyer=buyer, seller=seller, commodity=commodity,
                   broker=broker, office=office, status=status, uom=uom,
                   transport=transport, date_from=date_from,
                   date_to=date_to, ship_from=ship_from, ship_to=ship_to,
                   q=universal)

    any_filter = any(filters.values())
    if any_filter:
        contracts_list = q.order_by(Contract.date.desc()).all()
    else:
        contracts_list = q.order_by(Contract.date.desc()).limit(500).all()

    customers = Customer.query.filter_by(is_active=True).order_by(Customer.customer_name).all()
    return render_template('bulk_edit.html', contracts=contracts_list, customers=customers)


@app.route('/contracts')
@login_required
def contracts():
    q = Contract.query

    # Nationality-based access control (brokers also see contracts they're on)
    q = _apply_contract_access(q)

    # Text filters
    number = request.args.get('number', '').strip()
    buyer = request.args.get('buyer', '').strip()
    seller = request.args.get('seller', '').strip()
    commodity = request.args.get('commodity', '').strip()
    broker = request.args.get('broker', '').strip()
    office = request.args.get('office', '').strip()
    status = request.args.get('status', '').strip()
    uom = request.args.get('uom', '').strip()
    transport = request.args.get('transport', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    ship_from = request.args.get('ship_from', '').strip()
    ship_to = request.args.get('ship_to', '').strip()
    payment_from = request.args.get('payment_from', '').strip()
    payment_to = request.args.get('payment_to', '').strip()
    universal = request.args.get('q', '').strip()
    exact = request.args.get('exact', '').strip()

    if number:    q = q.filter(Contract.salesorder_number.ilike(f'%{number}%'))
    if buyer:     q = q.filter(Contract.cf_buyer.ilike(f'%{buyer}%'))
    if seller:    q = q.filter(Contract.customer_name.ilike(f'%{seller}%'))
    if commodity: q = q.filter(Contract.item_name.ilike(f'%{commodity}%'))
    if broker:    q = q.filter(Contract.salesperson_name.ilike(f'%{broker}%'))
    if office:    q = q.filter(Contract.location_name.ilike(f'%{office}%'))
    if uom:       q = q.filter(Contract.cf_uom.ilike(f'%{uom}%'))
    if transport: q = q.filter(Contract.cf_trnspname == transport)
    if date_from: q = q.filter(Contract.date >= date_from)
    if date_to:   q = q.filter(Contract.date <= date_to)
    if ship_from: q = q.filter(Contract.shipment_date >= ship_from)
    if ship_to:   q = q.filter(Contract.shipment_date <= ship_to)
    if payment_from: q = q.filter(Contract.paid_date >= payment_from)
    if payment_to:   q = q.filter(Contract.paid_date <= payment_to)
    # Status filtering mirrors the badge precedence in contracts.html:
    # Declined > Closed (paid) > underlying Zoho status. Without this, filtering
    # for e.g. "Open" returned rows the table then badged as Closed or Declined.
    from sqlalchemy import or_ as _or
    _not_declined = Contract.is_declined.isnot(True)
    _not_paid = _or(Contract.paid_status.is_(None), Contract.paid_status != 'paid')
    if status == 'declined':
        q = q.filter(Contract.is_declined == True)  # noqa
    elif status == 'closed':
        q = q.filter(_not_declined, Contract.paid_status == 'paid')
    elif status:
        q = q.filter(_not_declined, _not_paid, Contract.status == status)

    if universal:
        from sqlalchemy import or_
        _ucols = [Contract.salesorder_number, Contract.cf_buyer, Contract.customer_name,
                  Contract.item_name, Contract.salesperson_name, Contract.location_name,
                  Contract.cf_vessel_name, Contract.cf_customer_ref, Contract.buyer_reference,
                  Contract.reference_number, Contract.cf_co_broker, Contract.cf_split_broker,
                  Contract.cf_origin_location, Contract.cf_trnspname, Contract.cf_uom]
        if exact:
            q = q.filter(or_(*[c == universal for c in _ucols]))
        else:
            q = q.filter(or_(*[c.ilike(f'%{universal}%') for c in _ucols]))

    filters = dict(number=number, buyer=buyer, seller=seller, commodity=commodity,
                   broker=broker, office=office, status=status, uom=uom,
                   transport=transport, date_from=date_from,
                   date_to=date_to, ship_from=ship_from, ship_to=ship_to,
                   payment_from=payment_from, payment_to=payment_to,
                   q=universal)

    any_filter = any(filters.values())
    if any_filter:
        contracts_list = q.order_by(Contract.date.desc()).all()
    else:
        contracts_list = q.order_by(Contract.date.desc()).limit(500).all()

    return render_template('contracts.html', contracts=contracts_list, filters=filters, exact=exact,
                           limited=not any_filter)


@app.route('/tasks')
@login_required
def tasks_view():
    from sqlalchemy import func
    task_counts = db.session.query(
        Task.books_sales_order_id,
        func.count(Task.id).label('total'),
        func.sum(db.case((Task.status == 'pending', 1), else_=0)).label('pending')
    ).group_by(Task.books_sales_order_id).all()

    contract_map = {c.salesorder_id: c for c in Contract.query.all()}
    emp_map = {e.employee_id: e.employee_name for e in Employee.query.all()}

    # Nationality filter — same gating as the all-contracts view; brokers also see contracts they're on
    restrict = (not current_user.is_admin) and current_user.nationality
    own_ids = _contracts_user_is_on() if restrict else set()

    task_list = []
    for row in task_counts:
        contract = contract_map.get(row.books_sales_order_id)
        if restrict and not (contract and (
                contract.location_name == current_user.nationality
                or contract.salesorder_id in own_ids)):
            continue
        pic_id = contract.pic_employee_id if contract else ''
        task_list.append({
            'salesorder_id': row.books_sales_order_id,
            'salesorder_number': contract.salesorder_number if contract else row.books_sales_order_id,
            'location_name': contract.location_name if contract else '',
            'pic_id': pic_id or '',
            'pic_name': emp_map.get(pic_id, '') if pic_id else '',
            'total': row.total,
            'pending': row.pending
        })

    task_list.sort(key=lambda x: x['pending'], reverse=True)
    active_employees = Employee.query.filter_by(is_active=True).order_by(Employee.employee_name).all()
    return render_template('tasks.html', task_list=task_list, active_employees=active_employees)


@app.route('/dashboard')
@login_required
def dashboard():
    orders = Contract.query.order_by(Contract.date.desc()).all()
    return render_template('dashboard.html', contracts=orders)


# endregion


# region Write routes
@app.route('/employees/<employee_id>/office', methods=['POST'])
@login_required
def update_employee_office(employee_id):
    if not current_user.is_admin:
        return {'ok': False, 'error': 'Admin only'}, 403
    employee = Employee.query.get(employee_id)
    if not employee:
        return {'ok': False, 'error': 'Employee not found'}, 404
    data = request.get_json()
    employee.office = data.get('office', '').strip() or None
    db.session.commit()
    return {'ok': True}


@app.route('/employees/<employee_id>/active', methods=['POST'])
@login_required
def update_employee_active(employee_id):
    employee = Employee.query.get(employee_id)
    if not employee:
        return {'ok': False, 'error': 'Employee not found'}, 404
    data = request.get_json()
    employee.is_active = bool(data.get('is_active', True))
    db.session.commit()
    return {'ok': True}


@app.route('/contracts/export')
@login_required
def export_contracts():
    from datetime import date
    try:
        buffer = _export_contracts_xlsx()
    except ImportError:
        # openpyxl absent from the deployed environment - surface it instead of
        # returning a bare 500 with no explanation.
        flash('Excel export unavailable: openpyxl is not installed on the server. Use Export to CSV.', 'danger')
        return redirect('/contracts')
    except Exception as e:
        app.logger.exception('Excel export failed')
        flash(f'Excel export failed: {e}', 'danger')
        return redirect('/contracts')
    filename = f"contracts_{date.today().isoformat()}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/contracts/export-csv')
def export_contracts_csv():
    """Export all contracts as CSV for Power BI or direct download (public endpoint)."""
    import csv
    from datetime import date
    from io import StringIO

    q = Contract.query.order_by(Contract.date.desc())

    # CSV headers matching the Excel export
    headers = [
        'Contract #', 'Date', 'Status', 'Seller', 'Buyer',
        'Seller Ref #', 'Buyer Ref #', 'Commodity', 'Quantity', 'UOM',
        'Commodity Price', 'Commission Rate', 'Commission Total',
        'Transportation', 'Vessel Name', 'Ship Date Start', 'Ship Date End',
        'Co-Broker', 'Co-Brokerage Rate', 'Salesperson', 'Office',
        'Packing', 'Notes', 'Terms', 'Declined',
    ]

    # Build CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for c in q.yield_per(500):
        writer.writerow([
            c.salesorder_number or '',
            c.date or '',
            c.status or '',
            c.customer_name or '',
            c.cf_buyer or '',
            c.cf_customer_ref or '',
            c.buyer_reference or '',
            c.item_name or '',
            c.quantity or '',
            c.cf_uom or '',
            c.cf_item_contract_price or '',
            c.rate or '',
            round((c.rate or 0) * (c.quantity or 0), 2) or '',
            c.cf_trnspname or '',
            c.cf_vessel_name or '',
            c.shipment_date or '',
            c.cf_shipment_end_date or '',
            c.cf_co_broker or '',
            c.cf_co_brokerage_rate or '',
            c.salesperson_name or '',
            c.location_name or '',
            c.packing or '',
            c.notes or '',
            c.terms or '',
            c.is_declined or '',
        ])

    # Convert to bytes
    output.seek(0)
    from io import BytesIO
    buffer = BytesIO(output.getvalue().encode('utf-8'))
    buffer.seek(0)

    filename = f"contracts_{date.today().isoformat()}.csv"
    return send_file(
        buffer,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )


@app.route('/contracts/<salesorder_id>')
@login_required
def contract_detail(salesorder_id):
    c = Contract.query.get(salesorder_id)
    if not c:
        flash('Contract not found.', 'danger')
        return redirect('/contracts')

    contract = {
        'salesorder_id': c.salesorder_id,
        'salesorder_number': c.salesorder_number,
        'status': c.status,
        'paid_status': c.paid_status,
        'is_declined': c.is_declined,
        'date': c.date,
        'shipment_date': c.shipment_date,
        'customer_id': c.customer_id,
        'customer_name': c.customer_name,
        'salesperson_name': c.salesperson_name,
        'salesperson_id': c.salesperson_id,
        'location_id': c.location_id,
        'location_name': c.location_name,
        'notes': c.notes,
        'terms': c.terms,
        'line_items': [{
            'item_id': c.item_id,
            'name': c.item_name,
            'rate': c.rate,
            'quantity': c.quantity,
        }] if c.item_id else [],
        'custom': {
            'cf_buyer': c.cf_buyer,
            'cf_item_contract_price': c.cf_item_contract_price,
            'cf_trnspname': c.cf_trnspname,
            'cf_uom': c.cf_uom,
            'cf_customer_ref': c.cf_customer_ref,
            'cf_co_broker': c.cf_co_broker,
            'cf_co_brokerage_rate': c.cf_co_brokerage_rate if c.cf_co_brokerage_rate is not None else '',
            'cf_split_broker': c.cf_split_broker,
            'cf_split_percentage': c.cf_split_percentage,
            'cf_vessel_name': c.cf_vessel_name,
            'cf_shipment_end_date': c.cf_shipment_end_date,
        },
    }

    customers = Customer.query.order_by(Customer.customer_name).all()
    items = Item.query.order_by(Item.item_name).all()

    tasks = Task.query.filter_by(
        books_sales_order_id=salesorder_id
    ).join(TaskTemplate).order_by(TaskTemplate.order).all()

    shipments = Shipment.query.filter_by(books_sales_order_id=salesorder_id).all()
    commissions = BrokerCommission.query.filter_by(books_sales_order_id=salesorder_id).all()
    notes = ContractNote.query.filter_by(books_sales_order_id=salesorder_id).order_by(ContractNote.created_at).all()
    employees = Employee.query.all()
    employee_map = {e.employee_id: e.employee_name for e in employees}
    active_employees = [e for e in employees if e.is_active]

    salesperson_employee = next(
        (e for e in employees if e.salesperson_id == c.salesperson_id), None
    )

    # locations + active employees (with .office) drive the broker→office
    # auto-population on the edit page, mirroring the new-contract form.
    locations = zoho.get_locations()

    return render_template('contract_detail.html',
                           contract=contract, customers=customers, items=items,
                           tasks=tasks, meta=c, shipments=shipments,
                           commissions=commissions, notes=notes, employee_map=employee_map,
                           active_employees=active_employees, locations=locations,
                           salesperson_employee=salesperson_employee)


@app.route('/contracts/<salesorder_id>/tasks')
@login_required
def contract_tasks_json(salesorder_id):
    tasks = Task.query.filter_by(
        books_sales_order_id=salesorder_id
    ).join(TaskTemplate).order_by(TaskTemplate.order).all()
    return [{
        'id': t.id,
        'title': t.template.title,
        'description': t.template.description,
        'status': t.status,
        'field': t.template.books_field or '',
        'editable': (t.template.books_field or '') in INLINE_TASK_FIELDS,
        'completed_value': t.completed_value or '',
        'completed_at': t.completed_at.isoformat() if t.completed_at else None,
    } for t in tasks]


# endregion


# region Administrator/Security
@app.route('/admin/audit')
@login_required
@admin_required
def audit_log_view():
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(200).all()
    return render_template('admin_audit.html', logs=logs)


@app.before_request
def audit_log():
    if request.method != 'POST':
        return
    if not current_user.is_authenticated:
        return

    path = request.path
    contract_paths = ['/contracts/', '/bulk_edit/']
    if not any(path.startswith(p) for p in contract_paths):
        return

    parts = path.strip('/').split('/')
    salesorder_id = parts[1] if len(parts) >= 2 and parts[0] == 'contracts' else None

    log = AuditLog(
        user=current_user.email,
        method=request.method,
        path=path,
        salesorder_id=salesorder_id,
        timestamp=datetime.now(timezone.utc)
    )
    db.session.add(log)
    db.session.commit()


@app.route('/admin/templates')
@login_required
@admin_required
def admin_templates():
    templates = TaskTemplate.query.order_by(TaskTemplate.country, TaskTemplate.order).all()
    return render_template('admin_templates.html', templates=templates)


@app.route('/admin/templates/create', methods=['POST'])
@login_required
@admin_required
def create_template():
    order = request.form.get('order', '').strip()
    template = TaskTemplate(
        country=request.form.get('country'),
        order=int(order) if order else 0,
        title=request.form.get('title'),
        description=request.form.get('description'),
        books_field=request.form.get('books_field')
    )
    db.session.add(template)
    db.session.commit()
    return redirect('/admin/templates')


@app.route('/admin/templates/<int:template_id>/update', methods=['POST'])
@login_required
@admin_required
def update_template(template_id):
    template = TaskTemplate.query.get(template_id)
    if not template:
        flash('Template not found.', 'danger')
        return redirect('/admin/templates')
    template.country = request.form.get('country')
    template.order = int(request.form.get('order', 0))
    template.title = request.form.get('title')
    template.description = request.form.get('description')
    template.books_field = request.form.get('books_field')
    db.session.commit()
    return redirect('/admin/templates')


@app.route('/admin/templates/<int:template_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_template(template_id):
    template = TaskTemplate.query.get(template_id)
    if not template:
        flash('Template not found.', 'danger')
        return redirect('/admin/templates')
    db.session.delete(template)
    db.session.commit()
    return redirect('/admin/templates')


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.role, User.display_name).all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/users/create', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    email = request.form.get('email', '').strip().lower()
    display_name = request.form.get('display_name', '').strip()
    role = request.form.get('role', 'broker')
    if not email:
        flash('Email is required.', 'danger')
        return redirect('/admin/users')
    if User.query.filter_by(email=email).first():
        flash(f'{email} already exists.', 'warning')
        return redirect('/admin/users')
    nationality = request.form.get('nationality', '').strip() or None
    db.session.add(User(email=email, display_name=display_name, role=role, nationality=nationality))  # noqa
    db.session.commit()
    flash(f'User {email} added.', 'success')
    return redirect('/admin/users')


@app.route('/admin/users/<int:user_id>/role', methods=['POST'])
@login_required
@admin_required
def admin_update_user_role(user_id):
    user = User.query.get(user_id)
    if not user:
        flash('User not found.', 'danger')
        return redirect('/admin/users')
    new_role = request.form.get('role')
    if new_role not in ('admin', 'broker'):
        flash('Invalid role.', 'danger')
        return redirect('/admin/users')
    user.role = new_role
    db.session.commit()
    flash(f'{user.email} role updated to {new_role}.', 'success')
    return redirect('/admin/users')


@app.route('/admin/users/<int:user_id>/nationality', methods=['POST'])
@login_required
@admin_required
def admin_update_user_nationality(user_id):
    user = User.query.get(user_id)
    if not user:
        flash('User not found.', 'danger')
        return redirect('/admin/users')
    user.nationality = request.form.get('nationality', '').strip() or None
    db.session.commit()
    flash(f'{user.email} nationality updated.', 'success')
    return redirect('/admin/users')


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = User.query.get(user_id)
    if not user:
        flash('User not found.', 'danger')
        return redirect('/admin/users')
    if user.id == current_user.id:
        flash('Cannot delete your own account.', 'danger')
        return redirect('/admin/users')
    db.session.delete(user)
    db.session.commit()
    flash(f'{user.email} removed.', 'success')
    return redirect('/admin/users')


# endregion


# region Submission/Api Call
@app.route('/submit_contract', methods=['POST'])
@login_required
def submit_contract_route():
    form_data = request.form.to_dict(flat=True)

    buyer_id = form_data.get('buyer', '')
    buyer = Customer.query.get(buyer_id)
    form_data['buyer_name'] = buyer.customer_name if buyer else ''

    broker_employees = request.form.getlist('broker_employee[]')
    broker_percentages = request.form.getlist('broker_percentage[]')
    form_data['salesperson_employee_id'] = broker_employees[0] if len(broker_employees) > 0 else ''
    form_data['second_broker_employee_id'] = broker_employees[1] if len(broker_employees) > 1 else ''
    form_data['second_broker_percentage'] = broker_percentages[1] if len(broker_percentages) > 1 else ''

    booking_numbers = request.form.getlist('booking_number[]')
    shipment_quantities = request.form.getlist('shipment_quantity[]')
    booking_parts = []
    for b, q in zip(booking_numbers, shipment_quantities):
        if b.strip():
            booking_parts.append(f'{b}: {q}' if q.strip() else b)
    form_data['booking_numbers_concat'] = ', '.join(booking_parts)

    # Append any "request new" notes to delivery_notes
    request_notes = []
    if form_data.get('requested_item_name', '').strip():
        request_notes.append('[REQUEST NEW ITEM: ' + form_data['requested_item_name'].strip() + ']')

    buyer_legal = form_data.get('requested_buyer_legal_name', '').strip()
    buyer_address = form_data.get('requested_buyer_address', '').strip()
    if buyer_legal or buyer_address:
        buyer_note = '[REQUEST NEW BUYER: ' + (buyer_legal or '(no name)') + (
            f' — {buyer_address}' if buyer_address else '') + ']'
        request_notes.append(buyer_note)

    seller_legal = form_data.get('requested_seller_legal_name', '').strip()
    seller_address = form_data.get('requested_seller_address', '').strip()
    if seller_legal or seller_address:
        seller_note = '[REQUEST NEW SELLER: ' + (seller_legal or '(no name)') + (
            f' — {seller_address}' if seller_address else '') + ']'
        request_notes.append(seller_note)

    # Collect all delivery notes boxes
    delivery_notes_list = request.form.getlist('delivery_notes[]')
    delivery_notes_list = [note.strip() for note in delivery_notes_list if note.strip()]
    all_notes = request_notes + delivery_notes_list

    form_data['delivery_notes'] = chr(10).join(all_notes) if all_notes else ''

    result = zoho.submit_contract(form_data)
    if result.get('code', 0) == 0:
        salesorder_id = result['salesorder']['salesorder_id']
        salesorder_number = result['salesorder']['salesorder_number']

        zoho.upsert_contract_from_zoho(result['salesorder'])

        office = request.form.get('location_name', 'Head Office')
        generate_tasks(salesorder_id, country=office)

        meta = Contract.query.get(salesorder_id)
        meta.buyer_reference = request.form.get('buyer_reference', '').strip() or None
        meta.pic_employee_id = request.form.get('pic_employee_id') or None
        meta.packing = request.form.get('packing') or None
        meta.origin = request.form.get('origin', '').strip() or None
        meta.paid_date = request.form.get('paid_date', '').strip() or None

        shipment_quantities = request.form.getlist('shipment_quantity[]')
        for booking, qty in zip(booking_numbers, shipment_quantities):
            if booking or qty:
                db.session.add(Shipment(
                    books_sales_order_id=salesorder_id,
                    booking_number=booking,
                    quantity=float(qty) if qty else None
                ))
        db.session.commit()

        employee_ids = request.form.getlist('broker_employee[]')
        percentages = request.form.getlist('broker_percentage[]')
        amounts = request.form.getlist('broker_amount[]')
        for emp_id, pct, amt in zip(employee_ids, percentages, amounts):
            if emp_id and pct:
                db.session.add(BrokerCommission(
                    books_sales_order_id=salesorder_id,
                    employee_id=emp_id,
                    percentage=float(pct),
                    amount=float(amt) if amt else 0.0
                ))
        db.session.commit()

        log = AuditLog(
            user=current_user.email,
            method='POST',
            path='/submit_contract',
            salesorder_id=salesorder_id,
            timestamp=datetime.now(timezone.utc)
        )
        db.session.add(log)
        db.session.commit()
        return redirect(f'/submit_success/{salesorder_id}?number={salesorder_number}')
    else:
        flash(f"Submission failed: {result.get('message', 'Unknown error')}", 'danger')
        return redirect('/new_contract?restore=1')


def _apply_contract_field(salesorder_id, field, value):
    """Push a single field change to Zoho and refresh the local cache.
    Shared by /bulk_edit/<id> and the tasks page inline editor.
    Returns (True, None) on success or (False, message) on failure."""
    contract = zoho.get_sales_order(salesorder_id)
    form_data = zoho.contract_to_form_data(contract)
    form_data[field] = value

    salesperson_zoho_id = contract.get('salesperson_id', '')
    if salesperson_zoho_id:
        matching_emp = Employee.query.filter_by(salesperson_id=salesperson_zoho_id).first()
        form_data['salesperson_employee_id'] = matching_emp.employee_id if matching_emp else ''
    else:
        form_data['salesperson_employee_id'] = ''

    local_shipments = Shipment.query.filter_by(books_sales_order_id=salesorder_id).all()
    booking_parts = []
    for s in local_shipments:
        if s.booking_number:
            qty_str = str(int(s.quantity)) if s.quantity and s.quantity == int(s.quantity) else str(
                s.quantity) if s.quantity else ''
            booking_parts.append(f'{s.booking_number}: {qty_str}' if qty_str else s.booking_number)
    form_data['booking_numbers_concat'] = ', '.join(booking_parts)

    result = zoho.update_contract(salesorder_id, form_data)
    if result.get('code', 0) == 0:
        check_task_reactivity(salesorder_id, {field: value})
        zoho.upsert_contract_from_zoho(result['salesorder'])
        return True, None
    return False, result.get('message', 'Unknown error')


# books_field values the tasks page can edit inline. Excludes buyer, seller and
# commodity (each needs a lookup against another table, not a free-text box) and
# shipping_date (the form field naming is inverted - see the shipment date TODO).
INLINE_TASK_FIELDS = {
    'vessel_name', 'commodity_rate', 'transportation', 'delivery_notes',
    'contract_date', 'commission_rate', 'quantity', 'uom', 'terms',
    'co_broker_name', 'co_brokerage_rate', 'seller_reference',
}


@app.route('/bulk_edit/<salesorder_id>', methods=['POST'])
@login_required
def bulk_edit_save(salesorder_id):
    data = request.get_json()
    field = data.get('field')
    value = data.get('value')

    allowed_fields = {
        'quantity', 'commission_rate', 'seller_reference', 'shipping_date_end', 'contract_date',
        'uom', 'transportation', 'co_broker_name', 'vessel_name',
        'commodity_rate', 'co_brokerage_rate', 'delivery_notes', 'terms',
        'buyer_reference', 'packing', 'buyer', 'seller', 'paid_date',
    }
    if field not in allowed_fields:
        return {'ok': False, 'error': 'Invalid field'}, 400

    if field in ('buyer_reference', 'packing', 'paid_date'):
        meta = Contract.query.get(salesorder_id)
        if not meta:
            meta = Contract(salesorder_id=salesorder_id)
            db.session.add(meta)
        if field == 'buyer_reference':
            meta.buyer_reference = value.strip() or None
        elif field == 'packing':
            meta.packing = value.strip() or None
        elif field == 'paid_date':
            meta.paid_date = value.strip() or None
        db.session.commit()
        return {'ok': True}

    if field in ('buyer', 'seller'):
        customer = Customer.query.get(value)
        if not customer:
            return {'ok': False, 'error': 'Customer not found'}, 400
        contract = zoho.get_sales_order(salesorder_id)
        form_data = zoho.contract_to_form_data(contract)
        if field == 'buyer':
            form_data['buyer_name'] = customer.customer_name
        else:
            form_data['seller'] = customer.customer_id
        result = zoho.update_contract(salesorder_id, form_data)
        if result.get('code', 0) == 0:
            zoho.upsert_contract_from_zoho(result['salesorder'])
            return {'ok': True}
        else:
            return {'ok': False, 'error': result.get('message', 'Unknown error')}, 500

    ok, err = _apply_contract_field(salesorder_id, field, value)
    if ok:
        return {'ok': True}
    return {'ok': False, 'error': err}, 500


@app.route('/contracts/<salesorder_id>/update', methods=['POST'])
@login_required
def update_contract(salesorder_id):
    form_data = request.form.to_dict(flat=True)

    buyer_id = form_data.get('buyer', '')
    buyer = Customer.query.get(buyer_id)
    form_data['buyer_name'] = buyer.customer_name if buyer else ''

    broker_employees = request.form.getlist('broker_employee[]')
    broker_percentages = request.form.getlist('broker_percentage[]')
    form_data['salesperson_employee_id'] = broker_employees[0] if len(broker_employees) > 0 else ''
    form_data['second_broker_employee_id'] = broker_employees[1] if len(broker_employees) > 1 else ''
    form_data['second_broker_percentage'] = broker_percentages[1] if len(broker_percentages) > 1 else ''

    booking_numbers = request.form.getlist('booking_number[]')
    shipment_quantities_update = request.form.getlist('shipment_quantity[]')
    booking_parts_update = []
    for b, q in zip(booking_numbers, shipment_quantities_update):
        if b.strip():
            booking_parts_update.append(f'{b}: {q}' if q.strip() else b)
    form_data['booking_numbers_concat'] = ', '.join(booking_parts_update)

    # Collect all delivery notes boxes
    delivery_notes_list = request.form.getlist('delivery_notes[]')
    delivery_notes_list = [note.strip() for note in delivery_notes_list if note.strip()]

    form_data['delivery_notes'] = chr(10).join(delivery_notes_list) if delivery_notes_list else ''

    # Quern-local fields are persisted BEFORE the Zoho call and committed
    # independently. Zoho refuses edits to sales orders that have already been
    # invoiced, and these fields have zero presence in any Zoho payload, so a
    # rejected push must not take a local-only edit (notably paid_date) down
    # with it. Mirrors the local-only short-circuit already used in
    # /bulk_edit/<salesorder_id>.
    meta = Contract.query.get(salesorder_id)
    if meta:
        meta.buyer_reference = request.form.get('buyer_reference', '').strip() or None
        meta.pic_employee_id = request.form.get('pic_employee_id') or None
        meta.packing = request.form.get('packing') or None
        meta.origin = request.form.get('origin', '').strip() or None
        meta.paid_date = request.form.get('paid_date', '').strip() or None
        db.session.commit()

    result = zoho.update_contract(salesorder_id, form_data)
    if result.get('code', 0) == 0:
        check_task_reactivity(salesorder_id, form_data)

        zoho.upsert_contract_from_zoho(result['salesorder'])

        Shipment.query.filter_by(books_sales_order_id=salesorder_id).delete()
        shipment_quantities = request.form.getlist('shipment_quantity[]')
        for booking, qty in zip(booking_numbers, shipment_quantities):
            if booking or qty:
                db.session.add(Shipment(
                    books_sales_order_id=salesorder_id,
                    booking_number=booking,
                    quantity=float(qty) if qty else None
                ))
        db.session.commit()

        BrokerCommission.query.filter_by(books_sales_order_id=salesorder_id).delete()
        employee_ids = request.form.getlist('broker_employee[]')
        percentages = request.form.getlist('broker_percentage[]')
        amounts = request.form.getlist('broker_amount[]')
        for emp_id, pct, amt in zip(employee_ids, percentages, amounts):
            if emp_id and pct:
                db.session.add(BrokerCommission(
                    books_sales_order_id=salesorder_id,
                    employee_id=emp_id,
                    percentage=float(pct),
                    amount=float(amt) if amt else 0.0
                ))
        db.session.commit()

        salesorder_number = result['salesorder']['salesorder_number']
        return redirect(f'/submit_success/{salesorder_id}?number={salesorder_number}&action=updated')
    else:
        flash(f"Zoho update failed: {result.get('message', 'Unknown error')} "
              f"(Quern-only fields such as Paid Date were saved.)", 'danger')
        return redirect(request.referrer or '/contracts')


# endregion


# region User auth and login routes
@app.route('/login')
def login():
    from core.auth import build_auth_url
    return redirect(build_auth_url())


@app.route('/auth/callback')
def auth_callback():
    from core.auth import acquire_token_by_code

    if current_user.is_authenticated:
        return redirect('/')

    if 'error' in request.args:
        flash(f"Sign-in failed: {request.args.get('error_description', request.args.get('error'))}", 'danger')
        return redirect('/login')

    try:
        result = acquire_token_by_code(request.args)
    except ValueError:
        return redirect('/login')
    if 'error' in result:
        flash(f"Token exchange failed: {result.get('error_description', result.get('error'))}", 'danger')
        return redirect('/login')

    claims = result.get('id_token_claims', {})
    email = (claims.get('preferred_username') or claims.get('email') or '').strip().lower()
    if not email:
        flash('Microsoft account did not return an email.', 'danger')
        return redirect('/login')

    user = User.query.filter_by(email=email).first()
    if not user:
        return redirect(f'/unauthorized?email={email}')

    if not user.display_name and claims.get('name'):
        user.display_name = claims.get('name')
        db.session.commit()

    login_user(user)
    return redirect('/')


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect('/logged-out')


@app.route('/logged-out')
def logged_out():
    return render_template('logged_out.html')


@app.route('/unauthorized')
def unauthorized():
    email = request.args.get('email', '')
    return render_template('unauthorized.html', email=email)


# endregion


# region Sub-screens like submission success
@app.route('/submit_success/<salesorder_id>')
@login_required
def submit_success(salesorder_id):
    salesorder_number = request.args.get('number', 'Unknown')
    return render_template('submit_success.html', salesorder_id=salesorder_id, salesorder_number=salesorder_number)


# endregion


# region Sub-processes like confirming tasks or seeding tasks
@app.route('/seed_tasks')
def seed_tasks():
    if TaskTemplate.query.first():
        return "Templates already seeded, skipping."
    templates = [
        TaskTemplate(country='Boulder', order=1, title='Task 1 - Vessel Name',
                     description='Confirm the vessel name with the counterparty.', books_field='vessel_name'),
        TaskTemplate(country='Boulder', order=2, title='Task 2 - Commodity Price',
                     description='Confirm the agreed commodity price.', books_field='commodity_rate'),
        TaskTemplate(country='Boulder', order=3, title='Task 3 - Quantity',
                     description='Confirm the quantity with the seller.', books_field='quantity'),
        TaskTemplate(country='Boulder', order=4, title='Task 4 - UOM',
                     description='Confirm the unit of measure.', books_field='uom'),
        TaskTemplate(country='Boulder', order=5, title='Task 5 - Shipping Date',
                     description='Confirm the expected shipping date.', books_field='shipping_date'),
    ]
    for t in templates:
        db.session.add(t)
    db.session.commit()
    return f"Seeded {len(templates)} task templates"


@app.route('/tasks/<int:task_id>/confirm', methods=['POST'])
@login_required
def confirm_task(task_id):
    data = request.get_json()
    task = Task.query.get(task_id)
    if not task:
        return {'ok': False, 'error': 'Task not found'}, 404
    value = (data.get('completed_value') or '').strip()

    # Field-backed tasks send the field's own value (a vessel name, a date, a
    # customer id), not the literal string 'yes'. The old check only accepted
    # 'yes', so every field-backed task fell through to the pending branch and
    # never saved. Yes/no tasks have no books_field and keep the 'yes' check.
    if task.template is not None and task.template.books_field:
        done = bool(value)
    else:
        done = value.lower() == 'yes'

    if done:
        task.status = 'complete'
        task.completed_value = value
        task.completed_at = datetime.now(timezone.utc)
    else:
        task.status = 'pending'
        task.completed_value = None
        task.completed_at = None
    db.session.commit()
    # Return the persisted status so the client stops assuming success.
    return {'ok': True, 'status': task.status}


@app.route('/tasks/<int:task_id>/value', methods=['POST'])
@login_required
def task_submit_value(task_id):
    """Set a task's linked contract field straight from the tasks page dropdown,
    then mark the task complete. Only free-text/numeric fields are accepted;
    anything needing a lookup still has to go through the contract page."""
    task = Task.query.get(task_id)
    if not task:
        return {'ok': False, 'error': 'Task not found'}, 404

    field = task.template.books_field if task.template else None
    if not field:
        return {'ok': False, 'error': 'Task has no linked field'}, 400
    if field not in INLINE_TASK_FIELDS:
        return {'ok': False, 'error': f'{field} must be set on the contract page'}, 400

    value = (request.get_json() or {}).get('value', '')
    if isinstance(value, str):
        value = value.strip()
    if not value:
        return {'ok': False, 'error': 'A value is required'}, 400

    ok, err = _apply_contract_field(task.books_sales_order_id, field, value)
    if not ok:
        return {'ok': False, 'error': err}, 500

    task.status = 'complete'
    task.completed_value = value
    task.completed_at = datetime.now(timezone.utc)
    db.session.commit()
    return {'ok': True, 'status': task.status}


# endregion


# region Dev page stuff
@app.route('/debug/contracts_list')
@login_required
def debug_contracts_list():
    orders = get_sales_orders(page=1)
    return orders[0] if orders else {}


@app.route('/debug/customers')
@login_required
def debug_customers():
    token = zoho.get_access_token()
    import requests
    response = requests.get(
        "https://www.zohoapis.com/books/v3/contacts",
        headers={"Authorization": f"Zoho-oauthtoken {token}"},
        params={"organization_id": zoho.ORG_ID, "contact_type": "customer", "per_page": 1}
    )
    return response.json()


@app.route('/debug/items')
@login_required
def debug_items():
    token = zoho.get_access_token()
    import requests
    response = requests.get(
        "https://www.zohoapis.com/books/v3/items",
        headers={"Authorization": f"Zoho-oauthtoken {token}"},
        params={"organization_id": zoho.ORG_ID, "per_page": 200}
    )
    return response.json()


@app.route('/dev')
@login_required
def dev_panel():
    return render_template('dev.html')


@app.route('/init')
@login_required
@admin_required
def init_db():
    db.create_all()
    wipe_tasks()
    seed_tasks()
    create_user()
    return {'result': 'DB initialized. Now hit /init/items, /init/customers, /init/employees in order.'}


@app.route('/init/items')
@login_required
@admin_required
def init_items():
    sync_items()
    return {'result': 'Items synced.'}


@app.route('/init/customers')
@login_required
@admin_required
def init_customers():
    sync_customers()
    return {'result': 'Customers synced.'}


@app.route('/init/employees')
@login_required
@admin_required
def init_employees():
    sync_employees()
    return {'result': 'Employees synced.'}


@app.route('/init/customers/page')
@login_required
@admin_required
def init_customers_page():
    page = request.args.get('page', 1, type=int)
    result = sync_customers_page(page)
    return {'has_more': result['has_more'], 'count': result['count'], 'page': page}


@app.route('/init/contracts/page')
@login_required
@admin_required
def init_contracts_page():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', None, type=int)  # noqa
    result = sync_contracts_page(page, limit=limit)
    return {'has_more': result['has_more'], 'count': result['count'], 'page': page}


@app.route('/init/incremental/page')
@login_required
def init_incremental_page():
    page = request.args.get('page', 1, type=int)
    result = zoho.incremental_sync_page(page)
    return {'has_more': result['has_more'], 'count': result['count'], 'page': page}


@app.route('/dev/action/<action>', methods=['POST'])
@login_required
def dev_action(action):
    with app.app_context():
        if action == 'wipe_tasks':
            result = wipe_tasks()
        elif action == 'seed_tasks':
            result = seed_tasks()
        elif action == 'sync_items':
            sync_items()
            result = 'Items synced.'
        elif action == 'sync_customers':
            sync_customers()
            result = 'Customers synced.'
        elif action == 'sync_employees':
            sync_employees()
            result = 'Employees synced.'
        elif action == 'create_user':
            result = create_user()
        elif action == 'wipe_audit':
            result = wipe_audit()
        elif action == 'wipe_users':
            result = wipe_users()
        elif action == 'first_time_startup':
            first_time_startup()
            result = 'First time startup complete.'
        else:
            result = f'Unknown action: {action}'
    return {'result': str(result)}


@app.route('/dev/debug/<target>')
@login_required
def dev_debug(target):
    if target == 'first_contract':
        result = debug_first_contract()
    elif target == 'locations':
        result = debug_locations()
    elif target == 'reporting_tags':
        result = zoho.debug_reporting_tags()
    else:
        result = {'error': f'Unknown debug target: {target}'}
    return result


@app.route('/dev/scheduler/log')
@login_required
def dev_scheduler_log():
    """Get scheduler log messages."""
    return {'log': scheduler_log, 'running': scheduler.running}


@app.route('/dev/scheduler/start', methods=['POST'])
@login_required
def dev_scheduler_start():
    """Resume the scheduler (if paused)."""
    if scheduler.running:
        try:
            scheduler.resume()
            log_scheduler_message('Scheduler manually resumed')
        except:
            pass  # Already running
        return {'ok': True, 'message': 'Scheduler running', 'running': True}
    return {'ok': True, 'message': 'Scheduler already running', 'running': True}


@app.route('/dev/scheduler/stop', methods=['POST'])
@login_required
def dev_scheduler_stop():
    """Stop the scheduler."""
    if scheduler.running:
        scheduler.pause()
        log_scheduler_message('Scheduler manually paused')
        return {'ok': True, 'message': 'Scheduler paused', 'running': False}
    return {'ok': True, 'message': 'Scheduler already stopped', 'running': False}


@app.route('/dev/scheduler/clear-log', methods=['POST'])
@login_required
def dev_scheduler_clear_log():
    """Clear the scheduler log."""
    global scheduler_log
    scheduler_log = []
    return {'ok': True, 'message': 'Log cleared'}


@app.route('/dev/sync/items', methods=['POST'])
@login_required
def dev_sync_items():
    """Manually trigger items sync."""
    try:
        log_scheduler_message('Manual items sync triggered via dev panel')
        sync_items()
        log_scheduler_message('Manual items sync completed')
        return {'ok': True, 'message': 'Items sync completed'}
    except Exception as e:
        msg = f'Items sync error: {str(e)}'
        log_scheduler_message(msg)
        return {'ok': False, 'error': msg}


@app.route('/dev/sync/counterparties', methods=['POST'])
@login_required
def dev_sync_counterparties():
    """Manually trigger counterparties sync."""
    try:
        log_scheduler_message('Manual counterparties sync triggered via dev panel')
        sync_customers()
        log_scheduler_message('Manual counterparties sync completed')
        return {'ok': True, 'message': 'Counterparties sync completed'}
    except Exception as e:
        msg = f'Counterparties sync error: {str(e)}'
        log_scheduler_message(msg)
        return {'ok': False, 'error': msg}


@app.route('/dev/sync/contracts', methods=['POST'])
@login_required
def dev_sync_contracts():
    """Manually trigger contracts sync."""
    try:
        log_scheduler_message('Manual contracts sync triggered via dev panel')
        page = 1
        total = 0
        while True:
            result = sync_contracts_page(page)
            total += result['count']
            log_scheduler_message(f'  Synced page {page}: {result["count"]} contracts')
            if not result['has_more']:
                break
            page += 1
        log_scheduler_message(f'Manual contracts sync completed ({total} total)')
        return {'ok': True, 'message': f'Contracts sync completed ({total} contracts)'}
    except Exception as e:
        msg = f'Contracts sync error: {str(e)}'
        log_scheduler_message(msg)
        return {'ok': False, 'error': msg}


@app.route('/dev/export/powerbi', methods=['POST'])
@login_required
def dev_export_powerbi():
    """Manually trigger Power BI Excel export."""
    try:
        log_scheduler_message('Manual Power BI export triggered via dev panel')
        output_path, result = export_contracts_to_excel()
        if output_path:
            log_scheduler_message(f'Manual Power BI export completed: {result} contracts exported to {output_path}')
            return {'ok': True, 'message': f'Power BI export completed: {result} contracts exported', 'path': output_path}
        else:
            msg = f'Power BI export error: {result}'
            log_scheduler_message(msg)
            return {'ok': False, 'error': msg}
    except Exception as e:
        msg = f'Power BI export error: {str(e)}'
        log_scheduler_message(msg)
        return {'ok': False, 'error': msg}


@app.route('/dev/download/powerbi', methods=['GET'])
@login_required
def dev_download_powerbi():
    """Download the Power BI export CSV file."""
    try:
        export_path = '/mnt/quern-data/contracts_export.csv'
        if not os.path.exists(export_path):
            flash('Export file not found. Run the export first.', 'danger')
            return redirect('/dev')
        return send_file(export_path, as_attachment=True, download_name='contracts_export.csv', mimetype='text/csv')
    except Exception as e:
        flash(f'Download error: {str(e)}', 'danger')
        return redirect('/dev')


@app.route('/debug_first_contract')
@login_required
def debug_first_contract():
    orders = get_sales_orders(page=1)
    if not orders:
        return {'error': 'No contracts found'}
    first_id = orders[0]['salesorder_id']
    contract = zoho.get_sales_order(first_id)
    return contract


@app.route('/debug_locations')
@login_required
def debug_locations():
    return {'locations': zoho.get_locations()}


# endregion


# region Local contract data

@app.route('/contracts/<salesorder_id>/decline', methods=['POST'])
@login_required
def toggle_decline(salesorder_id):
    c = Contract.query.get(salesorder_id)
    if not c:
        return {'ok': False, 'error': 'Contract not found'}, 404

    was_declined = bool(c.is_declined)
    c.is_declined = not was_declined

    # When newly declining, record the reason as a contract note
    if not was_declined:
        data = request.get_json(silent=True) or {}
        reason = (data.get('reason') or '').strip()
        if reason:
            db.session.add(ContractNote(
                books_sales_order_id=salesorder_id,
                author=current_user.display_name or current_user.email,
                body=f'[DECLINED] {reason}'
            ))

    db.session.commit()
    return {'ok': True, 'is_declined': c.is_declined}


@app.route('/contracts/<salesorder_id>/delete', methods=['POST'])
@login_required
def delete_contract(salesorder_id):
    if not current_user.is_admin:
        return {'ok': False, 'error': 'Admin access required'}, 403

    c = Contract.query.get(salesorder_id)
    if not c:
        return {'ok': False, 'error': 'Contract not found'}, 404

    # Delete related records
    Shipment.query.filter_by(books_sales_order_id=salesorder_id).delete()
    BrokerCommission.query.filter_by(books_sales_order_id=salesorder_id).delete()
    ContractNote.query.filter_by(books_sales_order_id=salesorder_id).delete()
    Task.query.filter_by(books_sales_order_id=salesorder_id).delete()
    AuditLog.query.filter_by(salesorder_id=salesorder_id).delete()

    # Delete contract
    db.session.delete(c)
    db.session.commit()

    return {'ok': True}


@app.route('/contracts/<salesorder_id>/shipments/add', methods=['POST'])
@login_required
def add_shipment(salesorder_id):
    data = request.get_json()
    qty_val = data.get('quantity')
    shipment = Shipment(
        books_sales_order_id=salesorder_id,
        booking_number=data.get('booking_number', ''),
        quantity=float(qty_val) if qty_val else None
    )
    db.session.add(shipment)
    db.session.commit()
    return {'ok': True, 'id': shipment.id}


@app.route('/contracts/shipments/<int:shipment_id>/delete', methods=['POST'])
@login_required
def delete_shipment(shipment_id):
    shipment = Shipment.query.get(shipment_id)
    if not shipment:
        return {'ok': False}, 404
    db.session.delete(shipment)
    db.session.commit()
    return {'ok': True}


@app.route('/contracts/<salesorder_id>/pdf')
@login_required
def contract_pdf(salesorder_id):
    contract, data = _build_contract_data(salesorder_id)
    if not contract:
        flash('Contract not found.', 'danger')
        return redirect('/contracts')
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo_header.jpeg')
    buffer = generate_contract_pdf(data, logo_path)
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"{contract.salesorder_number or 'contract'}.pdf"
    )


@app.route('/contracts/<salesorder_id>/docx')
@login_required
def contract_docx(salesorder_id):
    contract, data = _build_contract_data(salesorder_id)
    if not contract:
        flash('Contract not found.', 'danger')
        return redirect('/contracts')
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo_header.jpeg')
    buffer = generate_contract_docx(data, logo_path)
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        as_attachment=True,
        download_name=f"{contract.salesorder_number or 'contract'}.docx"
    )


@app.route('/contracts/<salesorder_id>/notes/add', methods=['POST'])
@login_required
def add_note(salesorder_id):
    body = request.form.get('body', '').strip()
    if not body:
        flash('Note cannot be empty.', 'warning')
        return redirect(f'/contracts/{salesorder_id}')
    note = ContractNote(
        books_sales_order_id=salesorder_id,
        author=current_user.display_name or current_user.email,
        body=body
    )
    db.session.add(note)
    db.session.commit()
    return redirect(f'/contracts/{salesorder_id}#notes')


@app.route('/contracts/notes/<int:note_id>/delete', methods=['POST'])
@login_required
def delete_note(note_id):
    note = ContractNote.query.get(note_id)
    if not note:
        return {'ok': False}, 404
    salesorder_id = note.books_sales_order_id
    db.session.delete(note)
    db.session.commit()
    return redirect(f'/contracts/{salesorder_id}#notes')


# endregion


# region Functions
def first_time_startup():
    wipe_tasks()
    seed_tasks()
    sync_items()
    sync_customers()
    sync_employees()
    create_user()


def _export_contracts_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter

    q = Contract.query
    q = _apply_contract_access(q)
    q = q.order_by(Contract.date.desc())

    headers = [
        'Contract #', 'Date', 'Status', 'Seller', 'Buyer',
        'Seller Ref #', 'Buyer Ref #', 'Commodity', 'Quantity', 'UOM',
        'Commodity Price', 'Commission Rate', 'Commission Total',
        'Transportation', 'Vessel Name', 'Ship Date Start', 'Ship Date End',
        'Co-Broker', 'Co-Brokerage Rate', 'Salesperson', 'Office',
        'Packing', 'Notes', 'Terms', 'Declined',
    ]
    col_widths = [
        15, 12, 12, 28, 28, 18, 18, 25, 12, 12,
        16, 16, 16, 16, 20, 14, 14,
        20, 16, 20, 15,
        12, 40, 40, 10,
    ]
    assert len(col_widths) == len(headers), 'col_widths must match headers'

    # write_only mode streams rows to a temp file → bounded memory for large exports
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('Contracts')

    # Column widths must be set before any rows are appended in write-only mode
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e1e1c', end_color='1e1e1c', fill_type='solid')
    header_align = Alignment(horizontal='center')
    header_cells = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        header_cells.append(cell)
    ws.append(header_cells)

    # Stream contracts in batches so the full set is never materialized at once
    for c in q.yield_per(500):
        ws.append([
            c.salesorder_number or '',
            c.date or '',
            c.status or '',
            c.customer_name or '',
            c.cf_buyer or '',
            c.cf_customer_ref or '',
            c.buyer_reference or '',
            c.item_name or '',
            c.quantity,
            c.cf_uom or '',
            c.cf_item_contract_price,
            c.rate,
            round((c.rate or 0) * (c.quantity or 0), 2) or None,
            c.cf_trnspname or '',
            c.cf_vessel_name or '',
            c.shipment_date or '',
            c.cf_shipment_end_date or '',
            c.cf_co_broker or '',
            c.cf_co_brokerage_rate,
            c.salesperson_name or '',
            c.location_name or '',
            c.packing or '',
            c.notes or '',
            c.terms or '',
            'Yes' if c.is_declined else 'No',
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_contract_data(salesorder_id):
    """Shared data-assembly for PDF and DOCX contract downloads."""
    c = Contract.query.get(salesorder_id)
    if not c:
        return None, {}

    customers = {cu.customer_id: cu.customer_name for cu in Customer.query.all()}  # noqa
    items = {i.item_id: i for i in Item.query.all()}

    commodity_item = items.get(c.item_id or '')
    origin = commodity_item.origin if commodity_item and commodity_item.origin else ''

    return c, {
        'date': c.date or '',
        'contract_number': c.salesorder_number or '',
        'seller': c.customer_name or '',
        'buyer': c.cf_buyer or '',
        'quantity': c.quantity,
        'uom': c.cf_uom or '',
        'shipment': c.shipment_date or '',
        'price': c.cf_item_contract_price,
        'other_conditions': c.notes or '',
        'broker': c.salesperson_name or '',
        'origin': origin,
        'quality': '', 'grades': '', 'weights': '', 'governing_contract': '',
        'discount': '', 'moisture': '', 'damage': '', 'heat_damage': '',
        'foreign_materials': '', 'splits': '', 'payment': '', 'demurrage': '',
    }


@app.route('/debug/contract_tags/<salesorder_id>')
def debug_contract_tags(salesorder_id):
    from core.zoho import get_access_token
    import requests
    import os

    org_id = os.getenv('ZOHO_ORG_ID')
    access_token = get_access_token()
    api_response = requests.get(
        f"https://www.zohoapis.com/books/v3/salesorders/{salesorder_id}",
        headers={"Authorization": f"Zoho-oauthtoken {access_token}"},
        params={"organization_id": org_id}
    )
    return api_response.json()


# endregion


if __name__ == '__main__':
    with app.app_context():
        db.create_all()

    app.run(debug=True, host='0.0.0.0')